from openpyxl import load_workbook
import spacy
from spacy.lang import en
import gensim
from gensim import parsing
from collections import Counter



#EXTRACTING DATA FROM THE EXCEL FILE

wb = load_workbook(filename = 'Actions-Keywords.xlsx')
sheet = wb['Sheet1']

Key_Words=[]
row_count=[]
action=[]

kw_col='C'
rc_col='D'
ac_col='B'


row_num=2
max_rc=sheet.max_row
print("Row Count = " + str(max_rc))
for i in range(row_num,max_rc):
	Key_Words.append(sheet[kw_col+str(i)].value)
	row_count.append(sheet[rc_col+str(i)].value)
	action.append(sheet[ac_col+str(i)].value)


#STARTING DEPENDECY GRAPH PART

nlp = spacy.load("en")

myfile = open("Fs_2.txt", "r")

document = myfile.read()
document = document.lower()
document_stem = document

document_stem = gensim.parsing.stem_text(document_stem)

searched_word=[]
searched_keywords=[]
document = nlp(document)
document_stem = nlp(document_stem)
found=[]
found_stem=[]
counter=-1
dict1={}
previous=0
previous_stem=0
dict1_stem={}
list_action_count=[]
list_action_count_two=[]
counter_k=0

for i in Key_Words:
	counter+=1
	words=i.split(",")
	searched_keywords.append(words)
	for word in words:
		wor=word.split()
		for wo in wor:
			wo=wo.lower()
			print(" \n \n Looking for " + str(wo))
			found.append([sent for sent in document.sents if wo in sent.string.lower()])
			len_found = len(found)
			print(len_found)
			#print(found[len_found-1])
			if(found[len_found-1] != []):
				print("Found " + str(wo))
				if(dict1.get(wo) == 1):

					dict1[wo]+=1
				else:
					dict1[wo] = 1
				
			else:
				dict1[wo] = 0

			previous=len_found
			print(" \n \n \n ")


			found_stem.append([sent for sent in document_stem.sents if wo in sent.string.lower()])
			searched_word.append(wo)
			len_found_stem = len(found_stem)

			if(found_stem[len_found_stem-1] != []):
				if(dict1_stem.get(wo) == 1):
					dict1_stem[wo]+=1
				else:
					dict1_stem[wo] = 1
				
			else:
				dict1_stem[wo] = 0

			previous_stem=len_found_stem





'''
print("Without Stemming")
for found_1 in found:
	for sente in found_1:
		#sentence = hotel[0] 
		for word in sente:
			print(str(word).encode("utf-8"), ': ', str(list(word.children)).encode("utf-8"))

print("With Stemming")
#print(found_stem)
for found_stem_1 in found_stem:
	for stem_sent in found_stem_1:
		for word in stem_sent:
			print(str(word).encode("utf-8"), ': ', str(list(word.children)).encode("utf-8"))

'''
#print (str(found_stem).encode("utf-8"))
#print(searched_word)

print (	 "Stemming")
dict1_stem={ k:v for k, v in dict1_stem.items() if v }
print(dict1_stem)

print ("Without Stemming")
#print (str(found).encode("utf-8"))

dict1={ k:v for k, v in dict1.items() if v }

print(dict1)

#print(str(found).encode("utf-8"))
for found_1 in found:
	#print ("Found _1 = " + str(found_1))
	for sente in found_1:		
		#print("senetences")
		print("The sentence is \n ")
		print (str(sente).encode("utf-8"))
		for word in sente:			
			#print(str(word).encode("utf-8"), ': ', str(list(word.children)).encode("utf-8"))
			bach=list(word.children)
			
			for child in bach:
				#print("Parent: " + str(word) + "Child :")
				#print(str(child).encode('utf-8'))
				counter=-1
				
				for K in Key_Words:
					counter+=1
					#print(counter)
					if (str(word) in K):
						print(action[counter].encode('utf-8'))
						print( "One word Action Found")
						list_action_count.append(counter)

					if ((str(word) + " " + str(child) in K ) or (str(child) + " " + str(word) in K )):
						#print("Counter = "  + str(counter))
						print((str(child) + " " + str(word)))
						print("Found Action")
						#print(action[counter].encode('utf-8'))
						list_action_count_two.append(counter)
						print("\n")




#print(list_action_count)

c=Counter(list_action_count)
print(sorted(c))

print ("For two ")

#print(list_action_count_two)

c_two=Counter(list_action_count_two)
print(sorted(c_two))



print(" BLA BLA BLA ")



